"""
CRM Export Service - Service for exporting Call records data to XLSX
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))

from datetime import datetime, timedelta
from app.models import db
from app.models.crm_model import Contact, Call, Campaign, BlacklistEntry
from sqlalchemy import func, and_, or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import tempfile

class CRMExportService:
    """Service for exporting CRM Call records data to XLSX format"""
    
    @staticmethod
    def export_calls_data(campaign_id=None, ankieter_id=None, date_from=None, date_to=None, status_filter=None):
        """
        Export calls data to XLSX format
        
        Args:
            campaign_id: Filter by campaign ID
            ankieter_id: Filter by ankieter ID  
            date_from: Filter calls from date
            date_to: Filter calls to date
            status_filter: Filter by call status
            
        Returns:
            tuple: (file_path, filename)
        """
        # Build query
        query = Call.query
        
        if campaign_id:
            query = query.filter_by(campaign_id=campaign_id)
        
        if ankieter_id:
            query = query.filter_by(ankieter_id=ankieter_id)
            
        if status_filter:
            query = query.filter_by(status=status_filter)
            
        if date_from:
            # Convert date to datetime for comparison
            date_from_datetime = datetime.combine(date_from, datetime.min.time())
            query = query.filter(Call.call_date >= date_from_datetime)
            
        if date_to:
            # Convert date to datetime for comparison
            date_to_datetime = datetime.combine(date_to, datetime.max.time())
            query = query.filter(Call.call_date <= date_to_datetime)
        
        # Get all calls with relationships
        calls = query.join(Contact).join(Campaign, isouter=True).order_by(Call.call_date.desc()).all()
        
        if not calls:
            raise ValueError("Brak danych do eksportu dla wybranych filtrów")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rekordy Call"
        
        # Define headers
        headers = [
            'ID Rekordu',
            'ID Kontaktu', 
            'Nazwa Kontaktu',
            'Telefon',
            'Email',
            'Firma',
            'Kampania',
            'Ankieter',
            'Status Połączenia',
            'Klasyfikator',
            'Data Dzwonienia',
            'Godzina Dzwonienia',
            'Czas Trwania (sekundy)',
            'Czas Trwania (minuty)',
            'Czas Pracy na Rekordzie',
            'Notatki Agent',
            'Priorytet',
            'Typ Kolejki',
            'Status Kolejki',
            'Data Utworzenia',
            'Data Aktualizacji',
            'Zaplanowana Data',
            'Następne Połączenie',
            'Próby Połączeń',
            'Maksymalne Próby',
            'Zablokowany',
            'Aktywny',
            'Numer Telefonu (Call)',
            'Twilio SID',
            'Event ID',
            'Lead Zarejestrowany'
        ]
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # Add data rows
        row = 2
        for call in calls:
            # Calculate work time on record (time from creation to last update)
            work_time = ""
            if call.created_at and call.updated_at:
                time_diff = call.updated_at - call.created_at
                total_seconds = time_diff.total_seconds()
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
                work_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Format call date and time
            call_date = ""
            call_time = ""
            if call.call_date:
                call_date = call.call_date.strftime('%Y-%m-%d')
                call_time = call.call_date.strftime('%H:%M:%S')
            
            # Format duration
            duration_seconds = call.duration or call.duration_seconds or 0
            duration_minutes = round(duration_seconds / 60, 2) if duration_seconds else 0
            
            # Get next call date
            next_call = ""
            if call.next_call_date:
                next_call = call.next_call_date.strftime('%Y-%m-%d %H:%M:%S')
            
            # Get scheduled date
            scheduled_date = ""
            if call.scheduled_date:
                scheduled_date = call.scheduled_date.strftime('%Y-%m-%d %H:%M:%S')
            
            # Get creation and update dates
            created_at = call.created_at.strftime('%Y-%m-%d %H:%M:%S') if call.created_at else ""
            updated_at = call.updated_at.strftime('%Y-%m-%d %H:%M:%S') if call.updated_at else ""
            
            # Map status to Polish
            status_mapping = {
                'lead': 'Lead',
                'rejection': 'Odmowa',
                'callback': 'Callback',
                'no_answer': 'Nie odebrał',
                'busy': 'Zajęty',
                'wrong_number': 'Błędny numer'
            }
            
            # Map priority to Polish
            priority_mapping = {
                'high': 'Wysoki',
                'medium': 'Średni', 
                'low': 'Niski'
            }
            
            # Map queue type to Polish
            queue_type_mapping = {
                'new': 'Nowy',
                'callback': 'Callback',
                'retry': 'Ponowienie'
            }
            
            # Map queue status to Polish
            queue_status_mapping = {
                'pending': 'Oczekujący',
                'in_progress': 'W trakcie',
                'completed': 'Ukończony',
                'cancelled': 'Anulowany'
            }
            
            # Prepare row data
            row_data = [
                call.id,  # ID Rekordu
                call.contact_id,  # ID Kontaktu
                call.contact.name if call.contact else "",  # Nazwa Kontaktu
                call.contact.phone if call.contact else "",  # Telefon
                call.contact.email if call.contact else "",  # Email
                call.contact.company if call.contact else "",  # Firma
                call.campaign.name if call.campaign else "",  # Kampania
                f"{call.ankieter.first_name} ({call.ankieter.email})" if call.ankieter else "",  # Ankieter
                status_mapping.get(call.status, call.status),  # Status Połączenia
                call.status,  # Klasyfikator (raw status)
                call_date,  # Data Dzwonienia
                call_time,  # Godzina Dzwonienia
                duration_seconds,  # Czas Trwania (sekundy)
                duration_minutes,  # Czas Trwania (minuty)
                work_time,  # Czas Pracy na Rekordzie
                call.notes or "",  # Notatki Agent
                priority_mapping.get(call.priority, call.priority),  # Priorytet
                queue_type_mapping.get(call.queue_type, call.queue_type),  # Typ Kolejki
                queue_status_mapping.get(call.queue_status, call.queue_status),  # Status Kolejki
                created_at,  # Data Utworzenia
                updated_at,  # Data Aktualizacji
                scheduled_date,  # Zaplanowana Data
                next_call,  # Następne Połączenie
                call.contact.call_attempts if call.contact else 0,  # Próby Połączeń
                call.contact.max_call_attempts if call.contact else 0,  # Maksymalne Próby
                "Tak" if call.contact and call.contact.is_blacklisted else "Nie",  # Zablokowany
                "Tak" if call.contact and call.contact.is_active else "Nie",  # Aktywny
                call.phone_number or "",  # Numer Telefonu (Call)
                call.twilio_sid or "",  # Twilio SID
                call.event_id or "",  # Event ID
                "Tak" if call.is_lead_registered else "Nie"  # Lead Zarejestrowany
            ]
            
            # Add row to worksheet
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Format numbers
                if col in [1, 2, 13]:  # ID fields and duration seconds
                    cell.number_format = '0'
                elif col == 14:  # Duration minutes
                    cell.number_format = '0.00'
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set minimum and maximum width
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"crm_calls_export_{timestamp}.xlsx"
        
        return temp_file.name, filename
    
    @staticmethod
    def get_export_summary(campaign_id=None, ankieter_id=None, date_from=None, date_to=None, status_filter=None):
        """
        Get summary statistics for export
        
        Returns:
            dict: Summary statistics
        """
        # Build query
        query = Call.query
        
        if campaign_id:
            query = query.filter_by(campaign_id=campaign_id)
        
        if ankieter_id:
            query = query.filter_by(ankieter_id=ankieter_id)
            
        if status_filter:
            query = query.filter_by(status=status_filter)
            
        if date_from:
            # Convert date to datetime for comparison
            date_from_datetime = datetime.combine(date_from, datetime.min.time())
            query = query.filter(Call.call_date >= date_from_datetime)
            
        if date_to:
            # Convert date to datetime for comparison
            date_to_datetime = datetime.combine(date_to, datetime.max.time())
            query = query.filter(Call.call_date <= date_to_datetime)
        
        # Get statistics
        total_calls = query.count()
        
        # Status breakdown
        status_stats = {}
        for call in query.all():
            status = call.status
            status_stats[status] = status_stats.get(status, 0) + 1
        
        # Priority breakdown
        priority_stats = {}
        for call in query.all():
            priority = call.priority or 'unknown'
            priority_stats[priority] = priority_stats.get(priority, 0) + 1
        
        # Date range
        first_call = query.order_by(Call.call_date.asc()).first()
        last_call = query.order_by(Call.call_date.desc()).first()
        
        return {
            'total_calls': total_calls,
            'status_breakdown': status_stats,
            'priority_breakdown': priority_stats,
            'date_range': {
                'from': first_call.call_date.strftime('%Y-%m-%d') if first_call and first_call.call_date else None,
                'to': last_call.call_date.strftime('%Y-%m-%d') if last_call and last_call.call_date else None
            },
            'filters_applied': {
                'campaign_id': campaign_id,
                'ankieter_id': ankieter_id,
                'status_filter': status_filter,
                'date_from': date_from.strftime('%Y-%m-%d') if date_from else None,
                'date_to': date_to.strftime('%Y-%m-%d') if date_to else None
            }
        }
    
    @staticmethod
    def get_available_campaigns():
        """Get list of available campaigns for export"""
        campaigns = Campaign.query.filter_by(is_active=True).order_by(Campaign.name).all()
        return [{'id': c.id, 'name': c.name} for c in campaigns]
    
    @staticmethod
    def get_available_ankieters():
        """Get list of available ankieter for export"""
        from app.models import User
        ankieter = User.query.filter(
            User.account_type.in_(['ankieter', 'admin']),
            User.is_active == True
        ).order_by(User.first_name).all()
        return [{'id': a.id, 'username': f"{a.first_name} ({a.email})"} for a in ankieter]
    
    @staticmethod
    def get_available_statuses():
        """Get list of available call statuses"""
        from app.config.crm_config import CALL_STATUSES
        return [{'value': key, 'label': value} for key, value in CALL_STATUSES.items()]
